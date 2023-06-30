import openpyxl
import pandas as pd
import io
from lender_waf_request import *
from db_size import *
from s3_size import *
from Servicecost import *
from Border import *
import json
from SES import *
from configuration_file import *

# Generic Lender wise cost report file name and its sheet names
s3_bucket = generic_bucket_name
generic_file_path = generic_report_file_path
sheet_a = generic_Service_cost_sheet
sheet_b = generic_Lender_waf_sheet
sheet_c = generic_DB_sheet
sheet_d = generic_s3_sheet
sheet_e = generic_LenderNames_sheet

def s3_export(a, b):
    # Read the existing Excel file from S3
    obj = s3.get_object(Bucket=generic_bucket_name, Key=generic_file_path)
    file_content = obj['Body'].read()

    # Load the workbook from the file content
    workbook = openpyxl.load_workbook(io.BytesIO(file_content))

    # Select the target sheet or create a new one if it doesn't exist
    if b in workbook.sheetnames:
        sheet = workbook[b]
    else:
        pass

    # Convert the DataFrame to a list of lists
    data_values = a.values.tolist()

    # Insert the column headers to the sheet starting from the first row
    column_headers = list(a.columns)
    sheet.insert_rows(1)
    for col_num, value in enumerate(column_headers, start=1):
        sheet.cell(row=1, column=col_num).value = value

    # Insert the data rows to the sheet starting from the second row
    for row in reversed(data_values):
        sheet.insert_rows(2)
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=2, column=col_num).value = value

    # Save the updated workbook to a BytesIO object
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Upload the updated Excel file back to S3, replacing the existing file
    s3.upload_fileobj(output_stream, generic_bucket_name, generic_file_path)

# This Function will append Lender names to the Generic Lender cost report.
def lenders_append(sheet_e):
    # Calling the function from the db-size.py Module
    lenders = lender_names()    
    # Add suffix to lender names
    lenders = lenders.str.strip() + '.cyncsoftware.com'
    # Read the existing Excel file from S3
    obj = s3.get_object(Bucket=s3_bucket, Key=generic_file_path)
    file_content = obj['Body'].read()

    # Load the workbook from the file content
    workbook = openpyxl.load_workbook(io.BytesIO(file_content))

    # Select the target sheet or create a new one if it doesn't exist
    if sheet_e in workbook.sheetnames:
        sheet = workbook[sheet_e]
    else:
        pass

    # Get the next available row in the target sheet
    next_row = 1

    # Append the column header
    column_header = "Lender Name"
    sheet.cell(row=next_row, column=1).value = column_header

    # Append the lender names to the sheet starting from the next row
    for lender in lenders:
        next_row += 1
        sheet.cell(row=next_row, column=1).value = lender

    # Save the updated workbook to a BytesIO object
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Upload the updated Excel file back to S3, replacing the existing file
    s3.upload_fileobj(output_stream, s3_bucket, generic_file_path)

# This Function will append service cost details to the Generic Lender cost report.
def service_cost_append(sheet_a):
    # Calling the function from the Servicecost.py Module
    service_cost = service()
    s3_export(service_cost, sheet_a)

# This Function will append Lender-Waf Percentage details to the Generic Lender cost report.
def Lender_request_append(sheet_b):
    # Calling the function from the lender_waf_request.py Module
    Lender_request = lender_request()
    s3_export(Lender_request, sheet_b)    

# This Function will append DB-size and Percentage details to the Generic Lender cost report.
def db_size_append(sheet_c):
    # Calling the function from the db_size.py Module
    database = db_size()
    s3_export(database, sheet_c)
 
# This Function will append s3-size and Percentage details to the Generic Lender cost report.
def s3_size_append(sheet_d):
    # Calling the function from the s3_size.py Module
    s3_list = s3_size()
    s3_export(s3_list, sheet_d)

def email():
    send_email_with_s3_attachment(sender, recipient, subject, body, bucket_name, file_name)

def execute_data_update():

    # Append Lender Name details
    lenders_append(sheet_e)
    print("Added Lender Names to Generic Lender Wise Cost Report")
    # time.sleep(10)

    # Append service cost details
    service_cost_append(sheet_a)
    print("Conversion of Service Cost and Updation to Generic Lender Wise Cost Report is Completed")

    # Append Lender-Waf Percentage details
    Lender_request_append(sheet_b)
    print("Conversion of Lender Waf Request to percentage and Updation to Generic Lender Wise Cost Report is Completed")

    # Append DB-size and Percentage details
    db_size_append(sheet_c)
    print("Conversion of DB Size to percentage and Updation to Generic Lender Wise Cost Report is Completed")

    # Append s3-size and Percentage details
    s3_size_append(sheet_d)
    print("Conversion of s3 Size to percentage and Updation to Generic Lender Wise Cost Report is Completed")

    sheet_names = [sheet_a, sheet_b, sheet_c, sheet_d, sheet_e]
    apply_borders(s3_bucket, generic_file_path, sheet_names)
    print("Borders applied to the Generic Lender Wise Report")


def lambda_handler(event, context):
    execute_data_update()
    email()
    return {
    'statusCode': 200,
    'body': json.dumps('Execution is completed from Lambda')
    }

    