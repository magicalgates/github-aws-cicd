import pandas as pd
import boto3
import io
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from configuration_file import *

s3_bucket = generic_bucket_name
generic_file_path = generic_report_file_path
sheet_a = generic_Service_cost_sheet
sheet_b = generic_Lender_waf_sheet
sheet_c = generic_DB_sheet
sheet_d = generic_s3_sheet
sheet_e = generic_LenderNames_sheet

def apply_border(ws):
    # Create a border style
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    # Apply borders to columns and rows till the data range
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.border = border

    return ws

def apply_borders(s3_bucket, generic_file_path, sheet_names):
    # Set up the S3 client
    s3 = boto3.client('s3')

    # Download the file from S3
    response = s3.get_object(Bucket=s3_bucket, Key=generic_file_path)
    file_content = response['Body'].read()

    # Load the workbook from the file content
    workbook = load_workbook(io.BytesIO(file_content))

    for sheet_name in sheet_names:
        # Select the worksheet
        ws = workbook[sheet_name]
        # Apply borders
        ws = apply_border(ws)

    # Save the updated workbook to a BytesIO object
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Upload the updated Excel file back to S3, replacing the existing file
    s3.upload_fileobj(output_stream, s3_bucket, generic_file_path)

if __name__ == '__main__':
    sheet_names = [sheet_a, sheet_b, sheet_c, sheet_d, sheet_e]
    apply_borders(s3_bucket, generic_file_path, sheet_names)
    print("Borders applied to the Excel file in the S3 bucket.")
